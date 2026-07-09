#!/usr/bin/env python3
"""write_to_excel.py — 把模型抽好的一篇论文数据写进主 Excel 的 4 个 sheet。

设计原则（重写自旧 yaml_to_excel.py）：
    脚本只做机械 I/O，零语义判断。模型在交接 json 里已经把每列对好了主 Excel 列名，
    脚本照着列名写进对应位置即可。

输入：一个 json 交接文件（模型 Step 2 产出），结构见 references/handoff-format.md
输出：直接写主 Excel（写前自动备份）

用法：
    py write_to_excel.py <handoff.json>
    py write_to_excel.py <handoff.json> --dry-run    # 只打印要写什么，不真写

4 个 sheet 写入规则：
    论文清单：每篇 1 行，append 到最后一个真数据行之后
    样本数据：N 行，append 到最后一个真数据行之后（找真数据行，不是 max_row）
    样本卡片：按 paper_code 在各卡 B 列查卡；写代号 + 变量表 E/F/G 列（避开公式/合并格）
    大矩阵：每篇 1 行，写到第一个空的硬值数据行（不碰 row 12+ 的公式联动行）
"""
from __future__ import annotations
import argparse
import io
import json
import re
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).resolve().parents[6]
MASTER = PROJECT_ROOT / "01_Projects" / "Review_离子凝胶" / "IONOGEL_dynamic (8).xlsx"

# 每个 sheet 的"主键列"（用来从下往上找最后一个真数据行）
SHEET_KEY_COL = {
    "论文清单": 2,   # Code
    "样本数据": 2,   # Paper_Code
}


def backup_master() -> Path:
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    dst = MASTER.with_name(f"{MASTER.stem}_BACKUP_{ts}.xlsx")
    shutil.copy2(MASTER, dst)
    return dst


def header_map(ws, header_row: int) -> dict[str, int]:
    """返回 {列名: 列号}，列名取 header_row 行。"""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            out[str(v).strip()] = c
    return out


def last_data_row(ws, key_col: int, header_row: int) -> int:
    """从下往上找 key_col 最后一个非空行（跳过 max_row 之后的空白格区）。"""
    for r in range(ws.max_row, header_row, -1):
        v = ws.cell(r, key_col).value
        if v is not None and str(v).strip():
            return r
    return header_row


def safe_set(ws, r: int, c: int, value) -> bool:
    """写单元格；若是合并格的非左上角则跳过（返回 False）。"""
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


# ---------------------------------------------------------------------------
# 1) 论文清单：append 1 行
# ---------------------------------------------------------------------------
def write_paper_list(wb, data: dict, dry: bool) -> str:
    ws = wb["论文清单"]
    hmap = header_map(ws, 1)
    row_data = dict(data.get("论文清单", {}))
    row_data["Code"] = data["paper_code"]

    write_row = last_data_row(ws, SHEET_KEY_COL["论文清单"], 1) + 1
    written, unmatched = [], []
    for col_name, val in row_data.items():
        if val in (None, ""):
            continue
        if col_name in hmap:
            if not dry:
                safe_set(ws, write_row, hmap[col_name], val)
            written.append(col_name)
        else:
            unmatched.append(col_name)
    msg = f"论文清单: row {write_row} 写 {len(written)} 列"
    if unmatched:
        msg += f"\n     ⚠️ 列名对不上主 Excel、数据没写: {unmatched}"
    return msg


# ---------------------------------------------------------------------------
# 2) 样本数据：append N 行
# ---------------------------------------------------------------------------
def write_sample_data(wb, data: dict, dry: bool) -> str:
    ws = wb["样本数据"]
    hmap = header_map(ws, 2)            # 列名在第 2 行
    rows = data.get("样本数据", [])
    if not rows:
        return "样本数据: 无行"

    start = last_data_row(ws, SHEET_KEY_COL["样本数据"], 2) + 1
    unmatched = set()
    for i, rd in enumerate(rows):
        rd = dict(rd)
        rd["Paper_Code"] = data["paper_code"]
        for col_name, val in rd.items():
            if val in (None, ""):
                continue
            if col_name in hmap:
                if not dry:
                    safe_set(ws, start + i, hmap[col_name], val)
            else:
                unmatched.add(col_name)
    msg = f"样本数据: row {start}-{start + len(rows) - 1} 写 {len(rows)} 行"
    if unmatched:
        msg += f"\n     ⚠️ 列名对不上主 Excel、数据没写: {sorted(unmatched)}"
    return msg


# ---------------------------------------------------------------------------
# 3) 样本卡片：按 paper_code 查卡 → 写代号 + 变量表 E/F/G
# ---------------------------------------------------------------------------
CARD_SIZE = 80   # 每张样本卡片占 80 行（含基础信息 + 7 类变量表）


def _shift_rows(formula: str, delta: int) -> str:
    """把公式里本表（无 sheet! 前缀）的行号 +delta；跨表引用（论文清单!...）保持不动。

    例：=IFERROR(VLOOKUP($B$562,论文清单!$B$2:$M$100,2,FALSE()),"") delta=80
    →  =IFERROR(VLOOKUP($B$642,论文清单!$B$2:$M$100,2,FALSE()),"")  （只动 $B$562）
    """
    protected: list[str] = []

    def _protect(m):
        protected.append(m.group(0))
        return f"\x00{len(protected) - 1}\x00"

    # 1. 先保护跨表引用：sheet名!单元格(区域)
    tmp = re.sub(
        r"[\w一-鿿]+!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?",
        _protect, formula,
    )
    # 2. 偏移本表行号
    tmp = re.sub(
        r"(\$?[A-Z]{1,3}\$?)(\d+)",
        lambda m: m.group(1) + str(int(m.group(2)) + delta), tmp,
    )
    # 3. 还原跨表引用
    for i, p in enumerate(protected):
        tmp = tmp.replace(f"\x00{i}\x00", p)
    return tmp


def ensure_blank_card(ws) -> int:
    """样本卡片末尾新建一张空白工作卡：复制用户预留的「空白卡片模板」块
    （标题 '【空白卡片模板…】'，结构+公式骨架+合并格都为扩容备好），行引用按
    偏移调整，标题改成 '论文卡片' 让 find_card 当空白卡用。返回新卡起始行。

    解决批量论文超过预留卡数的问题——卡用完就照模板长一张新的。
    """
    template = None
    card_starts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        s = str(v)
        if s.startswith("【空白卡片模板"):
            template = r
        elif s.startswith("论文卡片"):
            card_starts.append(r)
    if template is None:                       # 没有专用模板块 → 退回复制最后一张卡
        template = card_starts[-1]

    last_block = max(card_starts + [template])
    new_start = last_block + CARD_SIZE
    delta = new_start - template

    # 先清掉新区域可能的残留合并格（避免 dst 是只读 MergedCell）
    for m in list(ws.merged_cells.ranges):
        if new_start <= m.min_row <= new_start + CARD_SIZE - 1:
            ws.unmerge_cells(str(m))

    for off in range(CARD_SIZE):
        for c in range(1, 8):
            dst = ws.cell(new_start + off, c)
            if isinstance(dst, MergedCell):
                continue
            src = ws.cell(template + off, c)
            v = src.value
            dst.value = _shift_rows(v, delta) if isinstance(v, str) and v.startswith("=") else v
            if src.has_style:
                dst._style = copy(src._style)
        h = ws.row_dimensions[template + off].height
        if h is not None:
            ws.row_dimensions[new_start + off].height = h

    # 标题改成"论文卡片"（让 find_card 当空白卡），清代号 + 变量表 E·F·G
    ws.cell(new_start, 1).value = "论文卡片  "
    ws.cell(new_start + 1, 2).value = None
    for off in range(15, CARD_SIZE):
        for col in (5, 6, 7):
            d = ws.cell(new_start + off, col)
            if not isinstance(d, MergedCell):
                d.value = None

    # 复制合并格（偏移）
    for m in list(ws.merged_cells.ranges):
        if template <= m.min_row <= template + CARD_SIZE - 1:
            ws.merge_cells(
                start_row=m.min_row + delta, start_column=m.min_col,
                end_row=m.max_row + delta, end_column=m.max_col,
            )
    return new_start


def find_card(ws, paper_code: str) -> int | None:
    """找该论文该写哪张卡片，返回卡起始行（'论文卡片 ...' 那行）。

    策略（避开主 Excel 已有的卡片代号错位问题，绝不覆盖别人数据）：
    1. 标题精确匹配 'SS-7' 且 该卡代号(B)为空或正好==目标 → 用它（更新自己）
    2. 否则用第一张「标题空 + 代号空」的全新空白卡
    3. 都没有 → None（卡片用完，需用户加）
    """
    cards = []  # [(start_row, 标题里的代号)]
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).startswith("论文卡片"):
            cards.append((r, str(v).replace("论文卡片", "").strip()))

    # 1. 标题匹配 + 代号没被别人占
    for start, label in cards:
        if label == paper_code:
            code = ws.cell(start + 1, 2).value
            if code is None or not str(code).strip() or str(code).strip() == paper_code:
                return start
            break   # 标题对但代号被别的论文占了 → 不覆盖，转空白卡

    # 2. 第一张全空白卡
    for start, label in cards:
        if not label:
            code = ws.cell(start + 1, 2).value
            if code is None or not str(code).strip():
                return start
    return None


def write_card(wb, data: dict, dry: bool) -> str:
    ws = wb["样本卡片"]
    card = data.get("样本卡片", {})
    if not card:
        return "样本卡片: 无数据"

    start = find_card(ws, data["paper_code"])
    auto_made = False
    if start is None:
        if dry:
            return "样本卡片: 预留空白卡用完 → 真写时会自动新建一张（dry-run 不建）"
        start = ensure_blank_card(ws)
        auto_made = True

    # 写论文代号（B<start+1>）
    if not dry:
        safe_set(ws, start + 1, 2, data["paper_code"])

    # 扫卡片区（start 到下一张卡前）的变量行，A 列是 var_code → 写 E/F/G
    end = start + 80
    written_vars = []
    for r in range(start, min(end, ws.max_row + 1)):
        vc = ws.cell(r, 1).value
        if vc is None:
            continue
        vc = str(vc).strip()
        if vc in card:
            v = card[vc]
            if not dry:
                if v.get("取值序列") not in (None, ""):
                    safe_set(ws, r, 5, v["取值序列"])      # E 列
                if v.get("样本数") not in (None, ""):
                    safe_set(ws, r, 6, v["样本数"])        # F 列
                if v.get("备注") not in (None, ""):
                    safe_set(ws, r, 7, v["备注"])          # G 列
            written_vars.append(vc)
    made = "（自动新建空白卡）" if auto_made else ""
    return f"样本卡片: 卡起始 row {start}{made}, 代号 {data['paper_code']}, 写变量 {written_vars}"


# ---------------------------------------------------------------------------
# 4) 大矩阵：不写！它是全公式联动 sheet。
#    每行变量列 = SUMIFS(样本卡片对应卡变量表)，B列代号 = IF(样本卡片对应卡代号)。
#    填好样本卡片（write_card）后，大矩阵对应行的公式会自动汇总，手写反而破坏公式。
# ---------------------------------------------------------------------------
def write_matrix(wb, data: dict, dry: bool) -> str:
    ws = wb["大矩阵"]
    # 找哪一行的代号公式会联动到我们刚写的样本卡片（B 列公式里引用了对应卡的代号格）
    linked_row = None
    for r in range(3, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.startswith("=") and "样本卡片" in b:
            linked_row = linked_row or r       # 第一个公式联动行起点
    return ("大矩阵: 不手写（全公式联动）— 填好样本卡片后对应行自动汇总，"
            f"SS 新卡会落在公式行区（首个公式行 row {linked_row}）")


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("handoff", help="模型产出的 json 交接文件")
    ap.add_argument("--dry-run", action="store_true", help="只打印不真写")
    args = ap.parse_args()

    hp = Path(args.handoff)
    if not hp.is_absolute():
        hp = PROJECT_ROOT / hp
    data = json.loads(hp.read_text(encoding="utf-8"))

    if "paper_code" not in data:
        print("❌ 交接文件缺 paper_code")
        sys.exit(1)

    if not MASTER.exists():
        print(f"❌ 主 Excel 不存在: {MASTER}")
        sys.exit(1)

    print(f"📄 paper_code: {data['paper_code']}")
    print(f"📊 主 Excel: {MASTER.name}")
    if args.dry_run:
        print("🔍 DRY RUN（不真写）")
    print()

    if not args.dry_run:
        bk = backup_master()
        print(f"💾 已备份: {bk.name}")
        print()

    wb = load_workbook(MASTER)
    results = [
        write_paper_list(wb, data, args.dry_run),
        write_sample_data(wb, data, args.dry_run),
        write_card(wb, data, args.dry_run),
        write_matrix(wb, data, args.dry_run),
    ]
    for r in results:
        print(f"  {r}")

    if not args.dry_run:
        try:
            wb.save(MASTER)
        except PermissionError:
            print(f"\n❌ 主 Excel 被占用（Excel 开着？），请关掉再跑")
            sys.exit(1)
        print(f"\n✅ 已写入主 Excel")
    else:
        print(f"\n（dry-run 结束，未写）")


if __name__ == "__main__":
    main()
